import openpyxl
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from os.path import join, abspath

# calculating the Chézy roughness coefficient C value
# with using the trained ANN (adjusted matrices of weights),
# weight matrices are downloaded from text files,
# input arguments are loaded from a MS Excel file,
# the calculation results are stored in a MS Excel file

# defining a function to calculate the derivative of the activation function

def logistic(x):
    return 1.0 / (1 + np.exp(-x))

# open the specified file only to read the values
data_path = join('.',"Input.xlsx") 
data_path = abspath(data_path) # it is created a full path to open the file based on

# the directory in which the program is running
Input_Data = openpyxl.open(data_path, read_only=True, data_only=True)
sheet_data = Input_Data.worksheets[0]

max_row_data = sheet_data['A3'].value # number of input data examples to calculate the Chézy roughness coefficient C value

# network parameters corresponding to the input data
input_size = sheet_data['A5'].value   # number of inputs
hidden_size = sheet_data['A7'].value  # number of neurons in the hidden layer
output_size = sheet_data['A9'].value  # number of outputs

# definition and initialization of matrices of the ANN input values and 
# matrix of the Chézy roughness coefficient calculated values

len_ryadok = input_size

characteristics_riverbed = np.zeros(
    shape=(max_row_data, len_ryadok))  # input - hydromorphological channel characteristics

coef_C = np.zeros(shape=(max_row_data))  # output – the Chézy coefficient value

# loading input values
for i in range(max_row_data):
    for j in range(len_ryadok):
        characteristics_riverbed[i][j] = sheet_data[i + 2][j + 1].value

print('The input data for the ANN was downloaded from the file:')
print(data_path)
print('The ANN parameters:')
print('number of inputs  = ', input_size)
print('number of neurons in the hidden layer  = ', hidden_size)
print('number of outputs  = ', output_size)
print('Number of input examples:', max_row_data)

list_param=[] # list of descriptions of input parameters
print('The coefficient C/100 is being investigated', end=' ')
for j in range(len_ryadok):
    list_param.append(sheet_data[1][j+1].value)
    print(sheet_data[1][j+1].value, end=',')
print()

Input_Data.close() # closing of the MS Excel input data file

# reading of matrix of weight coefficients W_1 і W_2

data_path = join('.','Data',"weights_matrix_1.txt")
data_path = abspath(data_path)

f = open(data_path)
raw_matrix1 = f.readlines()
f.close()

data_path = join('.','Data',"weights_matrix_2.txt")
data_path = abspath(data_path)

f = open(data_path)
raw_matrix2 = f.readlines()
f.close()

# initialization of matrices W_1 і W_2
W_1 = np.zeros(
    shape=(input_size, hidden_size))
W_2 = np.zeros(shape=(hidden_size, output_size))

# filling of matrices W_1 і W_2
for i in range (len(raw_matrix1)):
    b=raw_matrix1[i]
    a=b.split(' ')
    # len(a) == len(raw_matrix2) - к-сть рядків матриці W_2 = к-сті стовбців W_1
    for j in range(len(raw_matrix2)):
        W_1[i][j]=float(a[j])

for i in range (len(raw_matrix2)):
    W_2[i]=float(raw_matrix2[i])

if (len(raw_matrix1)==input_size)and(len(raw_matrix2)==hidden_size):
    print('the input data correspond to the parameters of the ANN matrices of weights,')
    # calculation of the network outputs
    for i in range(max_row_data):
        # direct course of calculations
        layer_0 = characteristics_riverbed[i:i + 1]
        layer_1 = logistic(np.dot(layer_0, W_1))
        layer_2 = np.dot(layer_1, W_2)
        coef_C[i] = layer_2
        print('for a set of parameters  №',i+1,', it is calculated C/100 =', coef_C[i])

# creating a MS Excel file to write data to a spreadsheet
    Out_Data=Workbook()
    sheet_data=Out_Data.active
    sheet_data.title="output_data"

    # recording data in the table
    # column names are added
    
    sheet_data['A1'].value='j'
    sheet_data.cell(row=1, column=len_ryadok+2).value='C/100'
   
    for j in range(len_ryadok):
        sheet_data[1][j + 1].value=list_param[j]

    #recording the input data and the corresponding calculated values of C
    for i in range(max_row_data):
        sheet_data['A'+str(i+2)].value = i+1
        # sheet_data['E'+str(i+2)].value = coef_C[i]
        sheet_data.cell(row=i+2, column=len_ryadok + 2).value = coef_C[i]
        for j in range(len_ryadok):
            sheet_data[i + 2][j + 1].value = characteristics_riverbed[i][j]

    #formatting the output to the table
    for i in range(1, len_ryadok+2):
        zag = sheet_data.cell(row=1, column=i)
        zag.alignment = Alignment(horizontal='center')
        zag.font = Font(bold=True, italic=False, color='DC143C', size=12)

        zag = sheet_data.cell(row=1, column=len_ryadok + 2)
        zag.alignment = Alignment(horizontal='center')
        zag.font = Font(bold=True, italic=False, color='000000', size=12)

    for i in range(max_row_data+1):
        sheet_data.cell(row=i+1, column=len_ryadok+2).fill=PatternFill(
            fill_type='solid', start_color='90EE90', end_color='90EE90'
        )


    exfilename = join('.', ('Output.xlsx')) # creating (join from parts) file name for storage
    exfilename=abspath(exfilename) #creating a complete path to store the file

    Out_Data.save(exfilename)  # storing the file
    Out_Data.close()
    print("Saving the calculation results in the file: ")
    print(exfilename)

else:
    print('The input data do not correspond to the ANN matrices of weights.')
    print('Check that the number of inputs and neurons in the hidden layer match ')
    print(' to the parameters of the trained network in the input file:')
    print('the number of columns W_1 (the number of neurons in the hidden layer of the network) = ', len(raw_matrix2))
    print('the number of rows W_1 (the number of the ANN inputs) = ', len(raw_matrix1))
    print('the number of rows W_2 (the number of neurons in the hidden layer) = ', len(raw_matrix2))

input('Press any key to end the program.')
